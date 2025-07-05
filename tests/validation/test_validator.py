#!/usr/bin/env python3
"""
Test validator module for QAA reports
Parses generated Excel files and validates against expected results
This module is separate from report generation logic to ensure true testing
"""

import os
import json
from openpyxl import load_workbook
from typing import Dict, List, Union, Any

class QAATestValidator:
    """Validates QAA report Excel files against expected results"""
    
    def __init__(self, expectations_file: str = "/home/rakanlinux/coolProjects/WebAlumni/tests/validation/test_expectations.json"):
        """Initialize validator with expected results"""
        self.expectations_file = expectations_file
        self.expectations = self._load_expectations()
    
    def _load_expectations(self) -> Dict:
        """Load expected results from JSON file"""
        try:
            with open(self.expectations_file, 'r') as f:
                return json.load(f)
        except FileNotFoundError:
            # Return default expectations if file doesn't exist yet
            return {
                "detailed_mode": {
                    "2014-2015": {
                        "total_graduates": 251
                    }
                },
                "simple_mode": {
                    "2014-2015": {
                        "total_graduates": 251,
                        "gentlemen": 162,
                        "ladies": 89
                    }
                }
            }
    
    def validate_detailed_mode_excel(self, excel_file_path: str, test_year: str = "2014-2015") -> Dict:
        """
        Validate detailed mode Excel file with status breakdown
        For detailed mode with 'combine all', we need to find the Total column and sum it
        """
        results = {
            "mode": "detailed",
            "test_year": test_year,
            "passed": False,
            "details": [],
            "errors": []
        }
        
        try:
            if not os.path.exists(excel_file_path):
                results["errors"].append(f"Excel file not found: {excel_file_path}")
                return results
            
            workbook = load_workbook(excel_file_path, data_only=True)
            
            # Look for the main data sheet (could be named differently)
            # Common names might be "Combined Report", "QAA Report", or the first sheet
            data_sheet = None
            possible_sheet_names = ["Combined Report", "QAA Report", workbook.sheetnames[0]]
            
            for sheet_name in possible_sheet_names:
                if sheet_name in workbook.sheetnames:
                    data_sheet = workbook[sheet_name]
                    break
            
            if data_sheet is None:
                results["errors"].append(f"Could not find data sheet. Available sheets: {workbook.sheetnames}")
                return results
            
            # Find the "Total" column
            total_column_index = None
            header_row = 1  # Assume first row is header
            
            for col in range(1, data_sheet.max_column + 1):
                cell_value = data_sheet.cell(row=header_row, column=col).value
                if cell_value and str(cell_value).strip().lower() == "total":
                    total_column_index = col
                    break
            
            if total_column_index is None:
                results["errors"].append("Could not find 'Total' column in Excel sheet")
                return results
            
            # Look for the "Overall Total" row instead of summing all values
            total_graduates = 0
            for row in range(header_row + 1, data_sheet.max_row + 1):
                # Check the first column for "Overall Total" label
                label_cell = data_sheet.cell(row=row, column=1).value
                if label_cell and isinstance(label_cell, str) and "overall total" in label_cell.lower():
                    cell_value = data_sheet.cell(row=row, column=total_column_index).value
                    if cell_value and isinstance(cell_value, (int, float)):
                        total_graduates = int(cell_value)
                        break
            
            # Get expected results
            expected = self.expectations.get("detailed_mode", {}).get(test_year, {})
            expected_total = expected.get("total_graduates", 0)
            
            # Validate total graduates
            total_passed = total_graduates == expected_total
            if total_passed:
                results["details"].append(f"✅ Total graduates: {total_graduates} (expected: {expected_total})")
            else:
                results["details"].append(f"❌ Total graduates: {total_graduates} (expected: {expected_total})")
            
            results["actual_total"] = total_graduates
            results["expected_total"] = expected_total
            
            # NEW: Status breakdown validation (if expected data exists)
            status_passed = True
            if expected.get("employed_count") is not None:
                employed_count = self._extract_status_count(data_sheet, "Employed", total_column_index)
                unemployed_count = self._extract_status_count(data_sheet, "Unemployed", total_column_index)
                studying_count = self._extract_status_count(data_sheet, "Studying", total_column_index)
                
                expected_employed = expected.get("employed_count", 0)
                expected_unemployed = expected.get("unemployed_count", 0)
                expected_studying = expected.get("studying_count", 0)
                
                # Validate each status
                if employed_count == expected_employed:
                    results["details"].append(f"✅ Employed: {employed_count} (expected: {expected_employed})")
                else:
                    results["details"].append(f"❌ Employed: {employed_count} (expected: {expected_employed})")
                    status_passed = False
                
                if unemployed_count == expected_unemployed:
                    results["details"].append(f"✅ Unemployed: {unemployed_count} (expected: {expected_unemployed})")
                else:
                    results["details"].append(f"❌ Unemployed: {unemployed_count} (expected: {expected_unemployed})")
                    status_passed = False
                
                if studying_count == expected_studying:
                    results["details"].append(f"✅ Studying: {studying_count} (expected: {expected_studying})")
                else:
                    results["details"].append(f"❌ Studying: {studying_count} (expected: {expected_studying})")
                    status_passed = False
                
                # Store actual values
                results["actual_employed"] = employed_count
                results["actual_unemployed"] = unemployed_count
                results["actual_studying"] = studying_count
                results["expected_employed"] = expected_employed
                results["expected_unemployed"] = expected_unemployed
                results["expected_studying"] = expected_studying
            
            # Overall pass/fail
            results["passed"] = total_passed and status_passed
            
        except Exception as e:
            results["errors"].append(f"Error validating detailed mode Excel: {str(e)}")
        
        return results
    
    def validate_simple_mode_excel(self, excel_file_path: str, test_year: str = "2014-2015") -> Dict:
        """
        Validate simple mode Excel file
        Simple mode creates sheets by academic year with gender breakdown
        """
        results = {
            "mode": "simple",
            "test_year": test_year,
            "passed": False,
            "details": [],
            "errors": []
        }
        
        try:
            if not os.path.exists(excel_file_path):
                results["errors"].append(f"Excel file not found: {excel_file_path}")
                return results
            
            workbook = load_workbook(excel_file_path, data_only=True)
            
            # Look for the year sheet - could be "2014-2015", "2014_2015", etc.
            year_sheet = None
            possible_sheet_names = [
                test_year,
                test_year.replace("-", "_"),
                test_year.replace("-", ""),
                f"Year_{test_year}",
                f"Academic_{test_year}"
            ]
            
            for sheet_name in possible_sheet_names:
                if sheet_name in workbook.sheetnames:
                    year_sheet = workbook[sheet_name]
                    break
            
            if year_sheet is None:
                results["errors"].append(f"Could not find year sheet for {test_year}. Available sheets: {workbook.sheetnames}")
                return results
            
            # In Simple mode, we need to find gender totals
            # The structure typically has colleges listed with Gentlemen/Ladies columns
            gentlemen_total = 0
            ladies_total = 0
            
            # Look for "Gentlemen" and "Ladies" columns
            gentlemen_col = None
            ladies_col = None
            
            # Search through the first few rows to find headers
            for row in range(1, 6):  # Check first 5 rows for headers
                for col in range(1, year_sheet.max_column + 1):
                    cell_value = year_sheet.cell(row=row, column=col).value
                    if cell_value:
                        cell_text = str(cell_value).strip().lower()
                        if "gentlemen" in cell_text:
                            gentlemen_col = col
                        elif "ladies" in cell_text:
                            ladies_col = col
            
            if gentlemen_col is None or ladies_col is None:
                results["errors"].append(f"Could not find 'Gentlemen' or 'Ladies' columns. Searching for alternative patterns...")
                
                # Alternative: look for "Male"/"Female" columns
                for row in range(1, 6):
                    for col in range(1, year_sheet.max_column + 1):
                        cell_value = year_sheet.cell(row=row, column=col).value
                        if cell_value:
                            cell_text = str(cell_value).strip().lower()
                            if "male" in cell_text and "female" not in cell_text:
                                gentlemen_col = col
                            elif "female" in cell_text:
                                ladies_col = col
            
            if gentlemen_col is None or ladies_col is None:
                results["errors"].append("Could not find gender columns in Simple mode Excel")
                return results
            
            # Look for the TOTAL row instead of summing all values
            for row in range(1, year_sheet.max_row + 1):
                # Check the first column for "TOTAL" label
                label_cell = year_sheet.cell(row=row, column=1).value
                if label_cell and isinstance(label_cell, str) and label_cell.strip().upper() == "TOTAL":
                    gentlemen_value = year_sheet.cell(row=row, column=gentlemen_col).value
                    ladies_value = year_sheet.cell(row=row, column=ladies_col).value
                    
                    if gentlemen_value and isinstance(gentlemen_value, (int, float)):
                        gentlemen_total = int(gentlemen_value)
                    
                    if ladies_value and isinstance(ladies_value, (int, float)):
                        ladies_total = int(ladies_value)
                    break
            
            total_graduates = gentlemen_total + ladies_total
            
            # Get expected results
            expected = self.expectations.get("simple_mode", {}).get(test_year, {})
            expected_total = expected.get("total_graduates", 0)
            expected_gentlemen = expected.get("gentlemen", 0)
            expected_ladies = expected.get("ladies", 0)
            
            # Validate each metric
            all_passed = True
            
            if total_graduates == expected_total:
                results["details"].append(f"✅ Total graduates: {total_graduates} (expected: {expected_total})")
            else:
                results["details"].append(f"❌ Total graduates: {total_graduates} (expected: {expected_total})")
                all_passed = False
            
            if gentlemen_total == expected_gentlemen:
                results["details"].append(f"✅ Gentlemen: {gentlemen_total} (expected: {expected_gentlemen})")
            else:
                results["details"].append(f"❌ Gentlemen: {gentlemen_total} (expected: {expected_gentlemen})")
                all_passed = False
            
            if ladies_total == expected_ladies:
                results["details"].append(f"✅ Ladies: {ladies_total} (expected: {expected_ladies})")
            else:
                results["details"].append(f"❌ Ladies: {ladies_total} (expected: {expected_ladies})")
                all_passed = False
            
            results["passed"] = all_passed
            results["actual_total"] = total_graduates
            results["actual_gentlemen"] = gentlemen_total
            results["actual_ladies"] = ladies_total
            results["expected_total"] = expected_total
            results["expected_gentlemen"] = expected_gentlemen
            results["expected_ladies"] = expected_ladies
            
        except Exception as e:
            results["errors"].append(f"Error validating simple mode Excel: {str(e)}")
        
        return results
    
    def _extract_status_count(self, sheet, status_name: str, total_column_index: int) -> int:
        """Extract count for specific employment status from Excel sheet"""
        try:
            # In detailed mode Excel, there are TWO sections:
            # 1. Individual status columns (Business owner, Employed, New graduate, etc.)
            # 2. Employment stats columns (aggregated Employed, Unemployed, Studying)
            # 
            # We need to read from the EMPLOYMENT STATS section, which comes after empty columns
            
            # Find the employment stats section by looking for the aggregated columns
            # These come after empty column(s) and contain the properly calculated totals
            
            employment_stats_columns = {}
            
            # Scan all headers to find the employment stats section
            for col in range(1, sheet.max_column + 1):
                header_cell = sheet.cell(row=1, column=col).value
                if header_cell and isinstance(header_cell, str):
                    header_clean = header_cell.strip()
                    
                    # Look for the employment stats columns (these are the aggregated ones)
                    if header_clean == "Employed":
                        # Check if this is in the employment stats section (not individual status section)
                        # Employment stats section comes after empty columns
                        is_employment_stats = False
                        
                        # Check if there are empty columns before this one
                        for check_col in range(max(1, col - 3), col):
                            check_header = sheet.cell(row=1, column=check_col).value
                            if not check_header or str(check_header).strip() == "":
                                is_employment_stats = True
                                break
                        
                        if is_employment_stats:
                            employment_stats_columns["Employed"] = col
                            
                    elif header_clean == "Unemployed":
                        # Similar check for unemployed in employment stats section
                        is_employment_stats = False
                        for check_col in range(max(1, col - 3), col):
                            check_header = sheet.cell(row=1, column=check_col).value
                            if not check_header or str(check_header).strip() == "":
                                is_employment_stats = True
                                break
                        
                        if is_employment_stats:
                            employment_stats_columns["Unemployed"] = col
                            
                    elif header_clean == "Studying":
                        # Similar check for studying in employment stats section
                        is_employment_stats = False
                        for check_col in range(max(1, col - 3), col):
                            check_header = sheet.cell(row=1, column=check_col).value
                            if not check_header or str(check_header).strip() == "":
                                is_employment_stats = True
                                break
                        
                        if is_employment_stats:
                            employment_stats_columns["Studying"] = col
            
            # Get the column for the requested status
            target_column = employment_stats_columns.get(status_name.title())
            
            if not target_column:
                return 0
            
            # Sum all values in this column (excluding header and "Overall Total" row)
            total_count = 0
            for row in range(2, sheet.max_row + 1):  # Skip header row
                # Skip the "Overall Total" row to avoid double counting
                first_cell = sheet.cell(row=row, column=1).value
                if first_cell and isinstance(first_cell, str) and "overall total" in first_cell.lower():
                    continue
                    
                cell_value = sheet.cell(row=row, column=target_column).value
                if cell_value and isinstance(cell_value, (int, float)):
                    total_count += int(cell_value)
            
            return total_count
            
        except Exception as e:
            return 0  # Return 0 on error
    
    def run_multi_year_test_suite(self, test_years: List[str], file_paths: Dict[str, Dict[str, str]]) -> Dict:
        """
        Run tests across multiple years
        
        Args:
            test_years: ["2014-2015", "2015-2016", "2016-2017"]
            file_paths: {
                "2014-2015": {"simple": "path1", "detailed": "path2"},
                "2015-2016": {"simple": "path3", "detailed": "path4"},
                # ...
            }
        """
        results = {
            "overall_passed": False,
            "years_tested": test_years,
            "year_results": {},
            "summary": []
        }
        
        all_passed = True
        for year in test_years:
            if year not in file_paths:
                results["summary"].append(f"❌ No file paths provided for year {year}")
                all_passed = False
                continue
                
            year_result = self.run_full_test_suite(
                file_paths[year]["simple"], 
                file_paths[year]["detailed"], 
                year
            )
            results["year_results"][year] = year_result
            
            if not year_result["overall_passed"]:
                all_passed = False
        
        results["overall_passed"] = all_passed
        
        # Generate summary
        for year in test_years:
            if year in results["year_results"]:
                year_result = results["year_results"][year]
                if year_result["overall_passed"]:
                    results["summary"].append(f"✅ {year}: All tests passed")
                else:
                    results["summary"].append(f"❌ {year}: Some tests failed")
            else:
                results["summary"].append(f"❌ {year}: No results available")
        
        return results
    
    def run_full_test_suite(self, simple_mode_file: str, detailed_mode_file: str, test_year: str = "2014-2015") -> Dict:
        """
        Run complete test suite on both Simple and Detailed mode files
        """
        results = {
            "test_year": test_year,
            "overall_passed": False,
            "simple_mode": {},
            "detailed_mode": {},
            "summary": []
        }
        
        # Test Simple mode
        simple_results = self.validate_simple_mode_excel(simple_mode_file, test_year)
        results["simple_mode"] = simple_results
        
        # Test Detailed mode  
        detailed_results = self.validate_detailed_mode_excel(detailed_mode_file, test_year)
        results["detailed_mode"] = detailed_results
        
        # Overall summary
        simple_passed = simple_results.get("passed", False)
        detailed_passed = detailed_results.get("passed", False)
        
        results["overall_passed"] = simple_passed and detailed_passed
        
        if simple_passed:
            results["summary"].append("✅ Simple mode tests PASSED")
        else:
            results["summary"].append("❌ Simple mode tests FAILED")
            
        if detailed_passed:
            results["summary"].append("✅ Detailed mode tests PASSED")
        else:
            results["summary"].append("❌ Detailed mode tests FAILED")
        
        return results

def format_test_results_for_display(results: Dict) -> str:
    """Format test results for web interface display"""
    output = []
    output.append(f"🧪 QAA Report Test Results for {results.get('test_year', 'Unknown Year')}")
    output.append("=" * 50)
    
    # Simple mode results
    simple = results.get("simple_mode", {})
    output.append("\n📊 SIMPLE MODE:")
    if simple.get("errors"):
        for error in simple["errors"]:
            output.append(f"❌ ERROR: {error}")
    else:
        for detail in simple.get("details", []):
            output.append(f"   {detail}")
    
    # Detailed mode results
    detailed = results.get("detailed_mode", {})
    output.append("\n📈 DETAILED MODE:")
    if detailed.get("errors"):
        for error in detailed["errors"]:
            output.append(f"❌ ERROR: {error}")
    else:
        for detail in detailed.get("details", []):
            output.append(f"   {detail}")
    
    # Overall summary
    output.append("\n🎯 SUMMARY:")
    for summary_item in results.get("summary", []):
        output.append(f"   {summary_item}")
    
    if results.get("overall_passed", False):
        output.append("\n🎉 ALL TESTS PASSED! 🎉")
    else:
        output.append("\n⚠️  SOME TESTS FAILED ⚠️")
    
    return "\n".join(output)

def format_multi_year_test_results_for_display(results: Dict) -> str:
    """Format multi-year test results for web interface display"""
    output = []
    years_tested = results.get("years_tested", [])
    output.append(f"🧪 Multi-Year QAA Report Test Results")
    output.append(f"Years tested: {', '.join(years_tested)}")
    output.append("=" * 60)
    
    # Results for each year
    for year in years_tested:
        year_result = results.get("year_results", {}).get(year, {})
        if not year_result:
            output.append(f"\n❌ {year}: No results available")
            continue
            
        output.append(f"\n🗓️  {year}:")
        output.append("-" * 20)
        
        # Simple mode for this year
        simple = year_result.get("simple_mode", {})
        output.append("📊 Simple Mode:")
        if simple.get("errors"):
            for error in simple["errors"]:
                output.append(f"   ❌ ERROR: {error}")
        else:
            for detail in simple.get("details", []):
                output.append(f"   {detail}")
        
        # Detailed mode for this year
        detailed = year_result.get("detailed_mode", {})
        output.append("📈 Detailed Mode:")
        if detailed.get("errors"):
            for error in detailed["errors"]:
                output.append(f"   ❌ ERROR: {error}")
        else:
            for detail in detailed.get("details", []):
                output.append(f"   {detail}")
        
        # Year summary
        if year_result.get("overall_passed", False):
            output.append(f"   ✅ {year} - ALL TESTS PASSED")
        else:
            output.append(f"   ❌ {year} - SOME TESTS FAILED")
    
    # Overall summary
    output.append(f"\n🎯 OVERALL SUMMARY:")
    for summary_item in results.get("summary", []):
        output.append(f"   {summary_item}")
    
    if results.get("overall_passed", False):
        output.append("\n🎉 ALL YEARS PASSED! 🎉")
    else:
        output.append("\n⚠️  SOME YEARS FAILED ⚠️")
    
    return "\n".join(output)

# Example usage and testing
if __name__ == "__main__":
    # For testing the validator independently
    validator = QAATestValidator()
    print("Test validator initialized successfully")
    print(f"Expectations loaded: {validator.expectations}")