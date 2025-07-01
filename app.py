# app.py - Flask application
import os
import pandas as pd
import json
import uuid
import re
import xlsxwriter
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_file, send_from_directory
from werkzeug.utils import secure_filename
from io import BytesIO
import shutil
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from copy import copy

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16 MB max upload
app.config['GENERATED_FILES'] = 'generated_files'

# Create upload and output directories if they don't exist
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['GENERATED_FILES'], exist_ok=True)

# Global session data storage - for demo purposes
# In production, you would use a proper database or session management
session_data = {}

# -----------------------------
# Define Constants
# -----------------------------
college_options = [
    "College of Engineering & Advan",
    "College of Business",
    "College of Science & General S",
    "College of Medicine",
    "College of Pharmacy"
]

# Expected Current Status values
expected_current_status = [
    "Employed",
    "Employed - add to list",
    "Business owner",
    "Training",
    "Do not contact",
    "Others",
    "Left the country",
    "Passed away",
    "Unemployed",
    "Studying",
    "New graduate"  # Added for banner integration
]

# Default fallback graduation years in case dynamic extraction fails
default_graduation_years = [
    "2010-2011 Spring", "2010-2011 Summer",
    "2011-2012 Spring", "2011-2012 Summer",
    "2012-2013 FALL", "2012-2013 Spring", "2012-2013 Summer",
    "2013-2014 FALL", "2013-2014 Spring", "2013-2014 Summer",
    "2014-2015 FALL", "2014-2015 Spring", "2014-2015 Summer",
    "2015-2016 FALL", "2015-2016 Spring", "2015-2016 Summer",
    "2016-2017 FALL", "2016-2017 Spring", "2016-2017 Summer",
    "2017-2018 FALL", "2017-2018 Spring", "2017-2018 Summer",
    "2018-2019 FALL", "2018-2019 Spring", "2018-2019 Summer",
    "2019-2020 FALL", "2019-2020 Spring", "2019-2020 Summer",
    "2020-2021 FALL", "2020-2021 Spring", "2020-2021 Summer",
    "2021-2022 FALL", "2021-2022 Spring", "2021-2022 Summer",
    "2022-2023 FALL", "2022-2023 Spring", "2022-2023 Summer",
    "2023-2024 Spring", "2023-2024 FALL", "2023-2024 Summer",
    "2024-2025 FALL"
]

# This will be populated dynamically from uploaded Excel files
graduation_years = []

# --------------------
# Helper Functions
# --------------------
def clean_status(status):
    if isinstance(status, str):
        return status.strip().lower().capitalize()
    return "Unknown"

def safe_get_column(df, column_name, default_value=""):
    """
    Safely get a column from dataframe, returning a default value series if the column doesn't exist
    """
    if column_name in df.columns:
        return df[column_name].fillna(default_value).astype(str)
    else:
        # Create an empty Series of the same length as the dataframe
        return pd.Series([default_value] * len(df), index=df.index)

def normalize_company_name(name):
    """
    Normalize company names with comprehensive empty value handling
    """
    # Handle None, NaN, and non-string types
    if pd.isna(name) or name is None:
        return "EMPTY (NULL)"

    if not isinstance(name, str):
        return "EMPTY (INVALID TYPE)"

    # Clean the string
    name = name.strip().upper()

    # Comprehensive empty/unknown patterns
    empty_patterns = {
        'COMPLETELY EMPTY': [
            '', '-', '.', ' ', '...', '*', '#', '//',
        ],
        'PLACEHOLDER': [
            '#N/A', 'N/A', 'NA', 'N.A.', 'N/A.', 'NONE', 'NIL',
            'NOT APPLICABLE', 'NOT AVAILABLE', 'UNKNOWN',
        ],
        'CONFIDENTIAL': [
            'CONFIDENTIAL', 'CONFIDENTIAL GOVERNMENT', 'GOVERNMENT',
            'GOVERNMENT SECTOR', 'CONFIDENTIAL (STEALTH MODE)',
            'CONFIDENTIAL ( STEALTH MODE )', 'CONFIDENTIAL COMPANY',
            'CANNOT DISCLOSE', 'UNDISCLOSED',
        ],
        'OTHERS': [
            'OTHERS', 'OTHER', 'MISC', 'MISCELLANEOUS', 'TBD',
            'TO BE DETERMINED', 'PENDING',
        ],
        'NOT WORKING': [
            'NOT WORKING', 'UNEMPLOYED', 'NO JOB', 'NO WORK',
            'LOOKING FOR JOB', 'SEEKING EMPLOYMENT',
        ]
    }

    # Check against empty patterns
    for category, patterns in empty_patterns.items():
        if name in patterns:
            return f"EMPTY ({category})"

    # Company aliases
    company_aliases = {
        # Banks
        'SNB': 'SAUDI NATIONAL BANK',
        'SAUDI NATIONAL BANK': 'SAUDI NATIONAL BANK',
        'SNB CAPITAL': 'SAUDI NATIONAL BANK',
        'THE SAUDI NATIONAL BANK': 'SAUDI NATIONAL BANK',
        'NCB': 'SAUDI NATIONAL BANK',
        'BSF': 'BANQUE SAUDI FRANSI',
        'BANQUE SAUDI FRANSI': 'BANQUE SAUDI FRANSI',
        'BANQUE SAUDI FRANSI CAPITAL': 'BANQUE SAUDI FRANSI',
        'FRANSI CAPITAL': 'BANQUE SAUDI FRANSI',
        'SAB': 'SAUDI BRITISH BANK',
        'SABB': 'SAUDI BRITISH BANK',
        'BANK SAB': 'SAUDI BRITISH BANK',

        # Government/PIF entities
        'PIF': 'PUBLIC INVESTMENT FUND',
        'PUBLIC INVESTMENT FUND - PIF': 'PUBLIC INVESTMENT FUND',
        'SAMA': 'SAUDI CENTRAL BANK',
        'SAUDI CENTRAL BANK - SAMA': 'SAUDI CENTRAL BANK',
        'SIDF': 'SAUDI INDUSTRIAL DEVELOPMENT FUND',
        'SAUDI INDUSTRIAL DEVELOPMENT FUND - SIDF': 'SAUDI INDUSTRIAL DEVELOPMENT FUND',

        # Consulting firms
        'BCG': 'BOSTON CONSULTING GROUP',
        'BOSTON CONSULTING GROUP (BCG)': 'BOSTON CONSULTING GROUP',
        'EY': 'ERNST & YOUNG',
        'ERNST & YOUNG (EY)': 'ERNST & YOUNG',
        'PWC': 'PRICEWATERHOUSECOOPERS',

        # Healthcare
        'KFSH&RC': 'KING FAISAL SPECIALIST HOSPITAL & RESEARCH CENTER',
        'KFSHRC': 'KING FAISAL SPECIALIST HOSPITAL & RESEARCH CENTER',
        'KFSH': 'KING FAISAL SPECIALIST HOSPITAL & RESEARCH CENTER',
        'KING FAISAL SPECIALIST HOSPITAL': 'KING FAISAL SPECIALIST HOSPITAL & RESEARCH CENTER',
        'HABIB': 'DR. SULAIMAN AL HABIB MEDICAL GROUP',
        'DR. SULAIMAN AL HABIB': 'DR. SULAIMAN AL HABIB MEDICAL GROUP',

        # Tech companies
        'STC': 'SAUDI TELECOM COMPANY',
        'SAUDI TELECOM': 'SAUDI TELECOM COMPANY',
        'HPE': 'HEWLETT PACKARD ENTERPRISE',
        'HEWLETT PACKARD ENTERPRISE - HPE': 'HEWLETT PACKARD ENTERPRISE',

        # Common typos/variations
        'ARAMCO': 'SAUDI ARAMCO',
        'SAUDI ARAMCO': 'SAUDI ARAMCO',
        'SABIC': 'SAUDI BASIC INDUSTRIES CORPORATION',
        'MINISTRY OF HEALTH': 'MINISTRY OF HEALTH',
        'MINISTRY OF HEALTH ': 'MINISTRY OF HEALTH',
    }

    # Try direct match first
    if name in company_aliases:
        return company_aliases[name]

    # Remove common words and standardize format
    remove_words = [
        'LTD', 'LIMITED', 'CORPORATION', 'CORP', 'INC', 'LLC', 'CO',
        'COMPANY', 'GROUP', 'HOLDING', 'HOLDINGS', 'INTERNATIONAL',
        'SAUDI ARABIA', 'KSA', 'MIDDLE EAST'
    ]

    for word in remove_words:
        name = name.replace(f' {word}', '')

    # Handle special cases with partial matches
    if 'NATIONAL GUARD' in name and 'HEALTH' in name:
        return 'MINISTRY OF NATIONAL GUARD HEALTH AFFAIRS'
    if 'KING FAHAD MEDICAL' in name:
        return 'KING FAHAD MEDICAL CITY'

    return name.strip()

def analyze_unknown_entries(df):
    """
    Analyze the distribution of unknown/empty workplace entries
    """
    # Create a copy of relevant columns
    analysis_df = df[['Current Workplace', 'Current Status', '_College', '_Year']].copy()

    # Categorize empty/unknown entries
    def categorize_empty(workplace, status):
        if not isinstance(workplace, str) or workplace.strip() == '':
            return "COMPLETELY EMPTY"
        workplace = workplace.upper().strip()
        if workplace in ['-', '#N/A', 'N/A', 'NA', 'NONE', 'UNKNOWN']:
            return f"PLACEHOLDER ({workplace})"
        if status.upper().strip() in ['UNEMPLOYED', 'STUDYING']:
            return f"STATUS: {status.upper().strip()}"
        return "OTHER"

    analysis_df['Empty_Category'] = analysis_df.apply(
        lambda x: categorize_empty(x['Current Workplace'], x['Current Status']),
        axis=1
    )

    return analysis_df['Empty_Category'].value_counts()

def is_high_position(position):
    """
    Determines if a job position is a high-level position.
    Returns a normalized position name if it's a high position, None otherwise.
    """
    if pd.isna(position) or position is None or not isinstance(position, str):
        return None
    
    # Clean and normalize the position string
    position = position.strip().upper()
    
    # Skip empty or placeholder values
    if not position or position in ['-', 'N/A', 'NA', 'NONE', 'NOT APPLICABLE', 'UNKNOWN']:
        return None
    
    # Define high position keywords with normalization mapping
    high_position_mapping = {
        # C-Suite positions
        'CEO': 'CHIEF EXECUTIVE OFFICER',
        'CHIEF EXECUTIVE OFFICER': 'CHIEF EXECUTIVE OFFICER',
        'PRESIDENT': 'PRESIDENT',
        'CFO': 'CHIEF FINANCIAL OFFICER',
        'CHIEF FINANCIAL OFFICER': 'CHIEF FINANCIAL OFFICER',
        'CTO': 'CHIEF TECHNOLOGY OFFICER',
        'CHIEF TECHNOLOGY OFFICER': 'CHIEF TECHNOLOGY OFFICER',
        'CIO': 'CHIEF INFORMATION OFFICER',
        'CHIEF INFORMATION OFFICER': 'CHIEF INFORMATION OFFICER',
        'COO': 'CHIEF OPERATING OFFICER',
        'CHIEF OPERATING OFFICER': 'CHIEF OPERATING OFFICER',
        'CMO': 'CHIEF MARKETING OFFICER',
        'CHIEF MARKETING OFFICER': 'CHIEF MARKETING OFFICER',
        'CHIEF': 'CHIEF',  # Generic chief
        
        # Director level
        'DIRECTOR': 'DIRECTOR',
        'EXECUTIVE DIRECTOR': 'EXECUTIVE DIRECTOR',
        'MANAGING DIRECTOR': 'MANAGING DIRECTOR',
        'BOARD MEMBER': 'BOARD MEMBER',
        
        # VP level
        'VP': 'VICE PRESIDENT',
        'VICE PRESIDENT': 'VICE PRESIDENT',
        'SVP': 'SENIOR VICE PRESIDENT',
        'SENIOR VICE PRESIDENT': 'SENIOR VICE PRESIDENT',
        'EVP': 'EXECUTIVE VICE PRESIDENT',
        'EXECUTIVE VICE PRESIDENT': 'EXECUTIVE VICE PRESIDENT',
        
        # Head positions
        'HEAD': 'HEAD',
        'DEPARTMENT HEAD': 'DEPARTMENT HEAD',
        'DIVISION HEAD': 'DIVISION HEAD',
        
        # Senior management
        'GENERAL MANAGER': 'GENERAL MANAGER',
        'PARTNER': 'PARTNER',
        'SENIOR MANAGER': 'SENIOR MANAGER',
        'PRINCIPAL': 'PRINCIPAL',
        
        # Founder positions
        'FOUNDER': 'FOUNDER',
        'CO-FOUNDER': 'CO-FOUNDER',
        'OWNER': 'OWNER'
    }
    
    # Check for exact matches first
    if position in high_position_mapping:
        return high_position_mapping[position]
    
    # Check for partial matches within the position title
    for keyword, normalized in high_position_mapping.items():
        # Use word boundaries to avoid partial word matches
        # e.g., 'DIRECTOR' should match 'IT DIRECTOR' but not 'DIRECTORY ADMINISTRATOR'
        pattern = r'\b' + re.escape(keyword) + r'\b'
        if re.search(pattern, position):
            # For common positions like HEAD and CHIEF, ensure it's not just part of another word
            if keyword in ['HEAD', 'CHIEF', 'OWNER'] and len(keyword) < 5:
                # Additional check to ensure it's actually a leadership position
                leadership_context = ['OF', 'DEPARTMENT', 'DIVISION', 'TEAM']
                if any(context in position for context in leadership_context):
                    return normalized
            else:
                return normalized
    
    return None

def get_workplace_statistics(df, colleges, years, degree_option, gender_option, nationality_option=None):
    """
    Generate workplace statistics from the given dataframe.
    """
    # Filter data based on selections
    mask = df["_College"].isin(colleges) & df["_Year"].isin(years)

    if degree_option == "bachelor":
        mask &= ~df["Student ID"].str.startswith("G", na=False)
    elif degree_option == "master":
        mask &= df["Student ID"].str.startswith("G", na=False)

    # Filter by gender if needed
    if gender_option.lower() != "all":
        mask &= (df["Gender"].str.strip().str.lower() == gender_option.lower())

    # Filter by nationality if Saudi is selected, only if the Nationality column exists
    if nationality_option and nationality_option.lower() == "saudi":
        if "Nationality" in df.columns:
            mask &= (df["Nationality"].str.strip() == "Saudi Arabia")
        else:
            print("Warning: Nationality column not found, nationality filter ignored")

    filtered_df = df[mask].copy()

    # Normalize company names
    filtered_df["Normalized_Workplace"] = filtered_df["Current Workplace"].apply(normalize_company_name)

    # Separate empty and valid entries
    empty_mask = filtered_df["Normalized_Workplace"].str.startswith("EMPTY", na=True)
    valid_df = filtered_df[~empty_mask]
    empty_df = filtered_df[empty_mask]

    # Get top employers (excluding empty values)
    top_employers = valid_df["Normalized_Workplace"].value_counts().head(10)

    # Get empty value statistics
    empty_stats = empty_df["Normalized_Workplace"].value_counts()

    # Get high positions (new implementation)
    # First, apply the high position detection function to create a new column
    filtered_df["High_Position"] = filtered_df["Current Position"].apply(is_high_position)
    
    # Filter only entries that have high positions (non-None values)
    high_position_df = filtered_df[filtered_df["High_Position"].notna()]
    
    # Count occurrences of each high position
    high_positions = high_position_df["High_Position"].value_counts().head(20)
    
    # Get original top positions for backward compatibility
    top_positions = filtered_df["Current Position"].value_counts().head(10)

    # Calculate nationality distribution (if Nationality column exists)
    if "Nationality" in df.columns:
        nationality_dist = filtered_df["Nationality"].value_counts().head(5)
    else:
        nationality_dist = pd.Series(dtype=int)  # Empty series

    # Calculate industry distribution (if Industry column exists)
    if "Industry" in df.columns:
        industry_dist = filtered_df["Industry"].value_counts().head(5)
    else:
        industry_dist = pd.Series(dtype=int)  # Empty series

    # Calculate employment type distribution (if Full Time or Part Time column exists)
    if "Full Time or Part Time" in df.columns:
        employment_type_dist = filtered_df["Full Time or Part Time"].value_counts()
    else:
        employment_type_dist = pd.Series(dtype=int)  # Empty series

    # Convert to dictionaries for JSON serialization
    return {
        "top_employers": top_employers.to_dict(),
        "empty_stats": empty_stats.to_dict(),
        "top_positions": top_positions.to_dict(),
        "high_positions": high_positions.to_dict(),  # New field for high positions
        "nationality_dist": nationality_dist.to_dict(),
        "industry_dist": industry_dist.to_dict(),
        "employment_type_dist": employment_type_dist.to_dict(),
        "total_alumni": len(filtered_df),
        "valid_entries": len(valid_df),
        "empty_entries": len(empty_df),
        "high_positions_count": len(high_position_df)  # Total count of alumni with high positions
    }

# --------------------
# Data Processing Functions
# --------------------
def load_excel_data(file_path, session_id):
    """
    Reads the Excel file and caches it for the session
    """
    try:
        # Read the Excel file
        df = pd.read_excel(file_path, dtype=str)
        df.columns = df.columns.str.strip()

        # Check if this is a Banner file by looking for Banner-specific columns
        is_banner_file = "Graduation Term" in df.columns and "Student Name" in df.columns

        if is_banner_file:
            # Required columns for Banner files
            required_columns = [
                "Student ID",
                "Student Name",
                "College",
                "Graduation Term",
                "Major",
                "Gender"
            ]
        else:
            # Required columns for Alumni files
            required_columns = [
                "College",
                "Year/Semester of Graduation",
                "Current Status",
                "Student ID",
                "Gender",
                "Major",
                "Current Workplace",
                "Current Position"
            ]

        # Optional columns that should be handled gracefully if missing
        optional_columns = [
            "Nationality",
            "Industry",
            "Full Time or Part Time",
            "Degree",
            "Personal Email",
            "Phone Number",
            "Minor",
            "Concentration",
            "GPA"
        ]

        # Check for required columns
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            return {"error": f"Missing required columns: {', '.join(missing_columns)}"}

        # Check for optional columns and warn if missing
        missing_optional = [col for col in optional_columns if col not in df.columns]
        warnings = []
        if missing_optional:
            warnings.append(f"The following optional columns are missing: {', '.join(missing_optional)}")

        # Handle missing values for critical columns
        for col in required_columns:
            df[col] = df[col].fillna("").astype(str).str.strip()

        # Handle optional columns if they exist
        for col in optional_columns:
            if col in df.columns:
                df[col] = df[col].fillna("").astype(str).str.strip()

        # Validate student IDs format (assuming they should be non-empty)
        invalid_ids = df[(df["Student ID"] != "") & (~df["Student ID"].str.match(r'^[0-9G]\d*$'))]
        if not invalid_ids.empty:
            warnings.append(f"Found {len(invalid_ids)} invalid Student IDs")

        # Create cleaned columns for fast filtering
        df["_College"] = df["College"].str.strip()
        if is_banner_file:
            df["_Year"] = df["Graduation Term"].str.strip()
        else:
            df["_Year"] = df["Year/Semester of Graduation"].str.strip()
            df["_CurrentStatus"] = df["Current Status"].str.strip().str.lower().str.capitalize()

        # Extract years from the file
        years_from_file, year_warnings = extract_graduation_years(file_path)
        
        # Add any warnings about year extraction
        warnings.extend(year_warnings)
        
        # Store the data in the session
        session_data[session_id] = {
            "data": df,
            "file_name": os.path.basename(file_path),
            "timestamp": datetime.now().isoformat(),
            "is_banner": is_banner_file,
            "graduation_years": years_from_file
        }
        
        # Update the global graduation_years if we found values
        global graduation_years
        if years_from_file and not is_banner_file:  # Don't use years from Banner files
            graduation_years = years_from_file

        # Validate graduation years
        years_to_validate = years_from_file if years_from_file else graduation_years
        if not years_to_validate:  # If still empty, use defaults
            years_to_validate = default_graduation_years
            
        invalid_years = df[~df["_Year"].isin(years_to_validate)]
        if not invalid_years.empty:
            # Get unique invalid years and convert to list
            unique_invalid_years = invalid_years["_Year"].unique().tolist()
            
            # Create a detailed warning message
            warning_msg = f"Found {len(invalid_years)} records with invalid graduation years: {unique_invalid_years}"
            
            # Optionally, you can also show which rows have these invalid years
            # This will help you locate them in the Excel file
            invalid_rows = invalid_years.index.tolist()
            warning_msg += f" (at rows: {invalid_rows[:10]}{'...' if len(invalid_rows) > 10 else ''})"
            
            warnings.append(warning_msg)

        # Validate colleges
        invalid_colleges = df[~df["_College"].isin(college_options)]
        if not invalid_colleges.empty:
            warnings.append(f"Found {len(invalid_colleges)} records with invalid colleges")
            
        # Validate Current Status values if not a Banner file
        if not is_banner_file and "Current Status" in df.columns:
            # Normalize the Current Status values for comparison
            df["_CurrentStatus"] = df["Current Status"].apply(clean_status)
            
            # Check for unexpected Current Status values
            normalized_expected_statuses = [clean_status(status) for status in expected_current_status]
            unexpected_statuses = df[~df["_CurrentStatus"].isin(normalized_expected_statuses)]
            
            if not unexpected_statuses.empty:
                unexpected_values = unexpected_statuses["_CurrentStatus"].unique().tolist()
                warnings.append(f"⚠️ ALERT: Found unexpected 'Current Status' values: {unexpected_values}")

        # Return some statistics about the loaded data
        stats = {
            "status": "success",
            "warnings": warnings,
            "record_count": len(df),
            "unique_colleges": df["_College"].nunique(),
            "college_distribution": df["_College"].value_counts().to_dict(),
            "year_range": {
                "min": df["_Year"].min(),
                "max": df["_Year"].max()
            },
            "gender_distribution": df["Gender"].value_counts().to_dict(),
            "is_banner": is_banner_file
        }

        return stats

    except Exception as e:
        return {"error": f"Error loading data: {str(e)}"}

def extract_graduation_years(file_path):
    """
    Extracts graduation years from an Excel file. 
    Looks for the "Year/Semester of Graduation" column to extract unique year values.
    """
    try:
        # Read the Excel file
        df = pd.read_excel(file_path, dtype=str)
        df.columns = df.columns.str.strip()
        
        if "Year/Semester of Graduation" in df.columns:
            # Extract unique values from the column
            years = df["Year/Semester of Graduation"].dropna().str.strip().unique().tolist()
            # Sort the years for better presentation
            years.sort()
            return years, []
        else:
            # Try to read from cell O1 (as mentioned in the task)
            try:
                wb = load_workbook(file_path, read_only=True)
                sheet = wb.active
                cell_value = sheet["O1"].value
                if cell_value and "Year/Semester of Graduation" in str(cell_value):
                    # Extract years from elsewhere in the Excel if possible
                    # This might need more specific logic based on the exact Excel structure
                    return default_graduation_years, ["Could not extract graduation years from O1, using defaults"]
                return default_graduation_years, ["Cell O1 does not contain Year/Semester of Graduation"]
            except Exception as e:
                return default_graduation_years, [f"Error reading cell O1: {str(e)}"]
    except Exception as e:
        return default_graduation_years, [f"Error extracting graduation years: {str(e)}"]

def create_gender_nationality_breakdown(filtered_df, writer, colleges):
    """
    Create a breakdown sheet showing employed vs unemployed stats by gender and nationality
    This adds a new sheet to the existing Excel writer
    """
    # Create a deep copy to avoid modifying the original dataframe
    breakdown_df = filtered_df.copy(deep=True)
    
    # Fix for missing columns if they don't exist
    if "Gender" not in breakdown_df.columns:
        print("Warning: Gender column not found in dataset")
        return
    if "Nationality" not in breakdown_df.columns:
        print("Warning: Nationality column not found in dataset")
        return
    if "Current Status" not in breakdown_df.columns:
        print("Warning: Current Status column not found in dataset")
        return
    
    # Create a sheet for gender/nationality breakdown
    sheet_name = "Gender_Nationality_Breakdown"
    
    # Initialize dataframes to store the results
    college_results = []
    
    # Process each college
    for college in colleges:
        college_df = breakdown_df[breakdown_df["College"].str.strip() == college].copy()
        if college_df.empty:
            continue
            
        # Determine employment status
        def get_employment_status(status):
            employed_statuses = [
                "Employed", "Employed - add to list", "Business owner", 
                "Training", "Do not contact", "Others", 
                "Left the country", "Passed away", "New graduate"
            ]
            if status in employed_statuses:
                return "Employed"
            elif status == "Unemployed":
                return "Unemployed"
            elif status == "Studying":
                return "Studying"
            else:
                return "Other"
        
        # Properly modify dataframe using .loc to avoid SettingWithCopyWarning        
        college_df.loc[:, "Employment Status"] = college_df["Current Status"].apply(get_employment_status)
        
        # Filter to just employed, unemployed, and studying and create a new dataframe
        emp_df = college_df[college_df["Employment Status"].isin(["Employed", "Unemployed", "Studying"])].copy()
        
        if emp_df.empty:
            continue
            
        # Create nationality category using .loc to avoid warning
        emp_df.loc[:, "Nationality Category"] = emp_df["Nationality"].fillna("").apply(
            lambda x: "Saudi" if x.strip() == "Saudi Arabia" else "Non-Saudi"
        )
        
        # Group by Gender, Nationality Category and Employment Status
        try:
            grouped = emp_df.groupby(["Gender", "Nationality Category", "Employment Status"]).size().unstack(
                fill_value=0
            ).reset_index()
            
            # Ensure all status columns exist
            if "Employed" not in grouped.columns:
                grouped["Employed"] = 0
            if "Unemployed" not in grouped.columns:
                grouped["Unemployed"] = 0
            if "Studying" not in grouped.columns:
                grouped["Studying"] = 0
                
            # Calculate totals and percentages
            grouped["Total"] = grouped["Employed"] + grouped["Unemployed"] + grouped["Studying"]
            grouped["Employed %"] = (grouped["Employed"] / grouped["Total"] * 100).fillna(0).round(2)
            grouped["Unemployed %"] = (grouped["Unemployed"] / grouped["Total"] * 100).fillna(0).round(2)
            grouped["Studying %"] = (grouped["Studying"] / grouped["Total"] * 100).fillna(0).round(2)
            
            # Add college name for multi-college display
            grouped["College"] = college
            
            college_results.append(grouped)
        except Exception as e:
            print(f"Error processing college {college}: {str(e)}")
            continue
    
    # Combine all results
    if college_results:
        try:
            all_results = pd.concat(college_results)
            
            # Reorganize for better readability
            all_results = all_results[["College", "Gender", "Nationality Category", 
                                    "Employed", "Unemployed", "Studying", "Total", 
                                    "Employed %", "Unemployed %", "Studying %"]]
            
            # Write to Excel
            all_results.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Format the sheet
            worksheet = writer.sheets[sheet_name]
            for idx, col in enumerate(all_results.columns):
                # Auto-adjust column width
                max_len = max(all_results[col].astype(str).map(len).max(), len(col)) + 2
                worksheet.set_column(idx, idx, max_len)
                
            # Add a title
            worksheet.write(all_results.shape[0] + 2, 0, "Gender and Nationality Employment Breakdown")
            worksheet.write(all_results.shape[0] + 3, 0, "Note: This breakdown shows employed vs. unemployed alumni by gender and nationality.")
            print("Successfully created gender/nationality breakdown sheet")
        except Exception as e:
            print(f"Error creating breakdown sheet: {str(e)}")

def process_qaa_report(session_id, colleges, years, degree_option, combine_all, combine_years, gender_option, nationality_option=None, mode_option="detailed"):
    """Generate QAA report based on given parameters"""
    try:
        if session_id not in session_data:
            return {"error": "No data found for your session. Please upload an Excel file first."}

        df = session_data[session_id]["data"].copy()
        
        # Add this line to normalize Current Status before grouping - THIS IS THE KEY FIX
        df["Current Status"] = df["Current Status"].apply(clean_status)
        
        # Filter based on selections - make a copy for the gender/nationality breakdown that doesn't filter by gender/nationality
        filtered_df_all = df[df["_College"].isin(colleges)]
        filtered_df_all = filtered_df_all[filtered_df_all["_Year"].isin(years)]
        
        # For Simple mode, ignore all filters except colleges and years
        if mode_option == "simple":
            filtered_df = filtered_df_all.copy()
        else:
            # Apply degree filter to the full dataset too
            if degree_option == "bachelor":
                filtered_df_all = filtered_df_all[~filtered_df_all["Student ID"].str.startswith("G", na=False)]
            elif degree_option == "master":
                filtered_df_all = filtered_df_all[filtered_df_all["Student ID"].str.startswith("G", na=False)]
                
            # Filter based on selections for the original report
            filtered_df = filtered_df_all.copy()
            
            # Filter by gender if needed
            if gender_option.lower() != "all":
                filtered_df = filtered_df[filtered_df["Gender"].str.strip().str.lower() == gender_option.lower()]

            # Filter by nationality if selected
            if nationality_option and nationality_option.lower() == "saudi":
                filtered_df = filtered_df[filtered_df["Nationality"].str.strip() == "Saudi Arabia"]
            elif nationality_option and nationality_option.lower() == "non-saudi":
                filtered_df = filtered_df[filtered_df["Nationality"].str.strip() != "Saudi Arabia"]

        if filtered_df.empty:
            return {"error": "No matching data found for the given filters."}

        # Generate a unique filename
        output_file = f"QAA_Report_{uuid.uuid4().hex[:8]}.xlsx"
        file_path = os.path.join(app.config['GENERATED_FILES'], output_file)
        
        # Handle Simple mode - completely different logic
        if mode_option == "simple":
            return process_simple_mode_report(filtered_df_all, colleges, years, file_path, output_file)
        
        # Use formatted majors if needed
        use_formatted_majors = combine_all or (degree_option == "all")
        if use_formatted_majors:
            filtered_df["Degree Type"] = filtered_df["Student ID"].apply(
                lambda sid: "Masters" if (isinstance(sid, str) and sid.strip().startswith("G")) else "Bachelors"
            )
            filtered_df["MajorFormatted"] = filtered_df["Major"].str.strip() + " - " + filtered_df["Degree Type"] + " - " + filtered_df["College"].str.strip()
            
            # Also add to the full dataset for the breakdown
            filtered_df_all["Degree Type"] = filtered_df_all["Student ID"].apply(
                lambda sid: "Masters" if (isinstance(sid, str) and sid.strip().startswith("G")) else "Bachelors"
            )
            filtered_df_all["MajorFormatted"] = filtered_df_all["Major"].str.strip() + " - " + filtered_df_all["Degree Type"] + " - " + filtered_df_all["College"].str.strip()

        with pd.ExcelWriter(file_path, engine="xlsxwriter") as writer:
            if combine_all:
                group_col = "MajorFormatted" if use_formatted_majors else "Major"
                grouped = filtered_df.groupby([group_col, "Current Status"]).size().unstack(fill_value=0)
                grouped["Total"] = grouped.sum(axis=1)
                
                # Calculate overall total for reference
                overall_total = grouped["Total"].sum()

                # Employment stats calculation - MATCH DESKTOP VERSION
                employment_stats = pd.DataFrame(index=grouped.index)
                employment_stats["Employed"] = (
                    grouped.get("Employed", 0) +
                    grouped.get("Employed - add to list", 0) +
                    grouped.get("Business owner", 0) +
                    grouped.get("Training", 0) +
                    grouped.get("Do not contact", 0) +
                    grouped.get("Others", 0) +
                    grouped.get("Left the country", 0) +
                    grouped.get("Passed away", 0) +
                    grouped.get("New graduate", 0)
                )
                employment_stats["Unemployed"] = grouped.get("Unemployed", 0)
                employment_stats["Studying"] = grouped.get("Studying", 0)

                # Calculate percentages using OVERALL TOTAL to match desktop version
                employment_stats["Employed Percentage"] = (employment_stats["Employed"] / overall_total * 100).fillna(0).round(2)
                employment_stats["Unemployed Percentage"] = (employment_stats["Unemployed"] / overall_total * 100).fillna(0).round(2)
                employment_stats["Studying Percentage"] = (employment_stats["Studying"] / overall_total * 100).fillna(0).round(2)

                empty_columns = pd.DataFrame(index=grouped.index, columns=["", " "])
                final_report = pd.concat([grouped, empty_columns, employment_stats], axis=1)
                
                # Calculate summary row MATCHING desktop version's approach
                summary_row = final_report.sum(axis=0).to_frame().T
                summary_row.index = ["Overall Total"]
                final_report = pd.concat([final_report, summary_row])
                
                final_report.to_excel(writer, sheet_name="Combined_Report")

                worksheet = writer.sheets["Combined_Report"]
                num_rows = final_report.shape[0]
                unique_years = sorted(set(filtered_df["Year/Semester of Graduation"].dropna().str.strip()))
                worksheet.write(num_rows + 5, 0, "Academic Years:")
                worksheet.write(num_rows + 5, 1, ", ".join(unique_years))

            elif combine_years:
                # Similar changes as above for the combine_years section
                college_abbreviations = {
                    "College of Engineering & Advan": "CoE",
                    "College of Business": "CoB",
                    "College of Medicine": "CoM",
                    "College of Pharmacy": "CoP",
                    "College of Science & General S": "CoS"
                }
                for college in colleges:
                    college_df = filtered_df[filtered_df["College"].str.strip() == college]
                    if college_df.empty:
                        continue

                    group_col = "MajorFormatted" if use_formatted_majors else "Major"
                    grouped = college_df.groupby([group_col, "Current Status"]).size().unstack(fill_value=0)
                    grouped["Total"] = grouped.sum(axis=1)
                    
                    overall_total = grouped["Total"].sum()

                    employment_stats = pd.DataFrame(index=grouped.index)
                    employment_stats["Employed"] = (
                        grouped.get("Employed", 0) +
                        grouped.get("Employed - add to list", 0) +
                        grouped.get("Business owner", 0) +
                        grouped.get("Training", 0) +
                        grouped.get("Do not contact", 0) +
                        grouped.get("Others", 0) +
                        grouped.get("Left the country", 0) +
                        grouped.get("Passed away", 0) +
                        grouped.get("New graduate", 0)
                    )
                    employment_stats["Unemployed"] = grouped.get("Unemployed", 0)
                    employment_stats["Studying"] = grouped.get("Studying", 0)

                    employment_stats["Employed Percentage"] = (employment_stats["Employed"] / overall_total * 100).fillna(0).round(2)
                    employment_stats["Unemployed Percentage"] = (employment_stats["Unemployed"] / overall_total * 100).fillna(0).round(2)
                    employment_stats["Studying Percentage"] = (employment_stats["Studying"] / overall_total * 100).fillna(0).round(2)

                    empty_columns = pd.DataFrame(index=grouped.index, columns=["", " "])
                    final_report = pd.concat([grouped, empty_columns, employment_stats], axis=1)
                    
                    summary_row = final_report.sum(axis=0).to_frame().T
                    summary_row.index = ["Overall Total"]
                    final_report = pd.concat([final_report, summary_row])

                    sheet_name = college_abbreviations.get(college.strip(), college[:25])
                    final_report.to_excel(writer, sheet_name=sheet_name)

                    worksheet = writer.sheets[sheet_name]
                    num_rows = final_report.shape[0]
                    unique_years = sorted(set(college_df["Year/Semester of Graduation"].dropna().str.strip()))
                    worksheet.write(num_rows + 5, 0, "Academic Years:")
                    worksheet.write(num_rows + 5, 1, ", ".join(unique_years))
            else:
                # Similar changes for the individual sheets by college and year
                college_abbreviations = {
                    "College of Engineering & Advan": "CoE",
                    "College of Business": "CoB",
                    "College of Science & General S": "CoS",
                    "College of Medicine": "CoM",
                    "College of Pharmacy": "CoP"
                }
                for college in colleges:
                    college_df = filtered_df[filtered_df["College"].str.strip() == college]
                    if college_df.empty:
                        continue
                    group_col = "MajorFormatted" if use_formatted_majors else "Major"
                    for year in years:
                        year_df = college_df[college_df["Year/Semester of Graduation"].str.strip() == year]
                        if year_df.empty:
                            continue

                        grouped = year_df.groupby([group_col, "Current Status"]).size().unstack(fill_value=0)
                        grouped["Total"] = grouped.sum(axis=1)
                        
                        overall_total = grouped["Total"].sum()

                        employment_stats = pd.DataFrame(index=grouped.index)
                        employment_stats["Employed"] = (
                            grouped.get("Employed", 0) +
                            grouped.get("Employed - add to list", 0) +
                            grouped.get("Business owner", 0) +
                            grouped.get("Training", 0) +
                            grouped.get("Do not contact", 0) +
                            grouped.get("Others", 0) +
                            grouped.get("Left the country", 0) +
                            grouped.get("Passed away", 0) +
                            grouped.get("New graduate", 0)
                        )
                        employment_stats["Unemployed"] = grouped.get("Unemployed", 0)
                        employment_stats["Studying"] = grouped.get("Studying", 0)

                        employment_stats["Employed Percentage"] = (employment_stats["Employed"] / overall_total * 100).fillna(0).round(2)
                        employment_stats["Unemployed Percentage"] = (employment_stats["Unemployed"] / overall_total * 100).fillna(0).round(2)
                        employment_stats["Studying Percentage"] = (employment_stats["Studying"] / overall_total * 100).fillna(0).round(2)

                        empty_columns = pd.DataFrame(index=grouped.index, columns=["", " "])
                        final_report = pd.concat([grouped, empty_columns, employment_stats], axis=1)
                        
                        summary_row = final_report.sum(axis=0).to_frame().T
                        summary_row.index = ["Overall Total"]
                        final_report = pd.concat([final_report, summary_row])

                        random_suffix = ''.join(re.findall(r'\w', uuid.uuid4().hex))[:3]
                        sheet_name = f"{college_abbreviations.get(college.strip(), college[:15])} - {year} - {random_suffix}"
                        if len(sheet_name) > 31:  # Excel sheet name length limit
                            sheet_name = sheet_name[:31]
                        final_report.to_excel(writer, sheet_name=sheet_name)

                        worksheet = writer.sheets[sheet_name]
                        num_rows = final_report.shape[0]
                        worksheet.write(num_rows + 5, 0, "Academic Year:")
                        worksheet.write(num_rows + 5, 1, year)

        # Add a separate with block to ensure the workbook is properly saved with our gender/nationality breakdown
        # Must save the workbook above first to avoid sharing violations

        # Now create a new excel writer to add our breakdown sheet
        breakdown_file = file_path
        try:
            # Use a new writer to add the gender/nationality breakdown sheet
            with pd.ExcelWriter(breakdown_file, engine="openpyxl", mode="a") as breakdown_writer:
                if not filtered_df_all.empty:
                    # Create the gender/nationality breakdown sheet after the regular reports are done
                    print("Adding gender/nationality breakdown sheet...")
                    create_gender_nationality_breakdown(filtered_df_all, breakdown_writer, colleges)
                    print("Breakdown sheet added successfully")
        except Exception as e:
            print(f"Failed to create gender/nationality breakdown: {str(e)}")
            
        return {
            "status": "success",
            "file": output_file
        }
            
    except Exception as e:
        return {"error": f"Error generating QAA report: {str(e)}"}


def process_simple_mode_report(filtered_df, colleges, years, file_path, output_file):
    """Generate Simple mode QAA report - just totals by college/year with gender breakdown"""
    try:
        # Helper function to extract academic year from graduation term
        def extract_academic_year(graduation_term):
            """Extract academic year (e.g., '2013-2014') from graduation term"""
            if pd.isna(graduation_term) or not isinstance(graduation_term, str):
                return "Unknown"
            
            # Look for patterns like "2013-2014", "2013-14", or just "2013"
            import re
            
            # Pattern for full academic year like "2013-2014"
            match = re.search(r'(\d{4})-(\d{4})', graduation_term)
            if match:
                return f"{match.group(1)}-{match.group(2)}"
            
            # Pattern for short academic year like "2013-14"
            match = re.search(r'(\d{4})-(\d{2})', graduation_term)
            if match:
                return f"{match.group(1)}-20{match.group(2)}"
            
            # Pattern for single year - assume it's the ending year
            match = re.search(r'(\d{4})', graduation_term)
            if match:
                year = int(match.group(1))
                prev_year = year - 1
                return f"{prev_year}-{year}"
            
            return "Unknown"
        
        # Normalize text data for case-insensitive comparisons
        filtered_df = filtered_df.copy()  # Make sure we're working with a copy
        filtered_df["Gender_Normalized"] = filtered_df["Gender"].str.strip().str.upper()
        filtered_df["College_Normalized"] = filtered_df["_College"].str.strip()
        
        # Add academic year column
        filtered_df["Academic_Year"] = filtered_df["_Year"].apply(extract_academic_year)
        
        # Get unique academic years, sorted
        academic_years = sorted(filtered_df["Academic_Year"].unique())
        
        with pd.ExcelWriter(file_path, engine="xlsxwriter") as writer:
            workbook = writer.book
            
            # Create formats
            header_format = workbook.add_format({
                'bold': True,
                'bg_color': '#4472C4',
                'font_color': 'white',
                'border': 1,
                'align': 'center'
            })
            cell_format = workbook.add_format({
                'border': 1,
                'align': 'left'
            })
            number_format = workbook.add_format({
                'border': 1,
                'align': 'center'
            })
            
            # Create one sheet per academic year
            for academic_year in academic_years:
                if academic_year == "Unknown":
                    continue  # Skip unknown years
                    
                # Filter data for this academic year
                year_df = filtered_df[filtered_df["Academic_Year"] == academic_year]
                
                if year_df.empty:
                    continue
                
                # Create summary by college and gender
                summary_data = []
                
                # Get all colleges that have data for this year
                colleges_with_data = year_df["College_Normalized"].unique()
                
                for college in colleges_with_data:
                    college_df = year_df[year_df["College_Normalized"] == college]
                    
                    # Total graduates for this college
                    total_graduates = len(college_df)
                    
                    # Gender breakdown (case-insensitive)
                    gender_counts = college_df["Gender_Normalized"].value_counts()
                    gentlemen = gender_counts.get("MALE", 0)
                    ladies = gender_counts.get("FEMALE", 0)
                    
                    summary_data.append({
                        "College": college,
                        "Total Graduates": total_graduates,
                        "Gentlemen": gentlemen,
                        "Ladies": ladies
                    })
                
                # Create DataFrame and write to Excel
                if summary_data:
                    summary_df = pd.DataFrame(summary_data)
                    
                    # Clean sheet name (Excel has 31 character limit)
                    sheet_name = academic_year.replace("-", "_")[:31]
                    
                    summary_df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Format the sheet
                    worksheet = writer.sheets[sheet_name]
                    
                    # Set column widths
                    worksheet.set_column('A:A', 35)  # College
                    worksheet.set_column('B:B', 15)  # Total Graduates
                    worksheet.set_column('C:C', 15)  # Gentlemen
                    worksheet.set_column('D:D', 15)  # Ladies
                    
                    # Add totals row
                    total_row = len(summary_df) + 2
                    worksheet.write(total_row, 0, "TOTAL", header_format)
                    worksheet.write(total_row, 1, summary_df["Total Graduates"].sum(), number_format)
                    worksheet.write(total_row, 2, summary_df["Gentlemen"].sum(), number_format)
                    worksheet.write(total_row, 3, summary_df["Ladies"].sum(), number_format)
            
            # Create "All years summary" tab
            create_all_years_summary_tab(filtered_df, writer, workbook, header_format, number_format)
        
        return {
            "status": "success",
            "file": output_file
        }
        
    except Exception as e:
        return {"error": f"Error generating Simple mode report: {str(e)}"}


def create_all_years_summary_tab(filtered_df, writer, workbook, header_format, number_format):
    """Create the 'All years summary' tab with college/gender and nationality breakdowns"""
    try:
        # Group data by college for overall summary
        summary_data = []
        
        # Get all colleges that have data across all years
        colleges_with_data = filtered_df["College_Normalized"].unique()
        
        for college in colleges_with_data:
            college_df = filtered_df[filtered_df["College_Normalized"] == college]
            
            # Total graduates for this college
            total_graduates = len(college_df)
            
            # Gender breakdown (case-insensitive)
            gender_counts = college_df["Gender_Normalized"].value_counts()
            gentlemen = gender_counts.get("MALE", 0)
            ladies = gender_counts.get("FEMALE", 0)
            
            # Nationality breakdown (case-insensitive)
            # Normalize nationality data
            nationality_normalized = college_df["Nationality"].fillna("").str.strip()
            saudi_count = len(nationality_normalized[nationality_normalized == "Saudi Arabia"])
            non_saudi_count = total_graduates - saudi_count
            
            summary_data.append({
                "College": college,
                "Total Graduates": total_graduates,
                "Ladies": ladies,
                "Gentlemen": gentlemen,
                "Saudi": saudi_count,
                "Non-Saudi": non_saudi_count
            })
        
        if not summary_data:
            return
            
        # Create DataFrame
        summary_df = pd.DataFrame(summary_data)
        
        # Write to "All years summary" sheet
        sheet_name = "All years summary"
        
        # Create worksheet manually for custom formatting
        worksheet = workbook.add_worksheet(sheet_name)
        
        # Define column positions based on the image layout
        # F-G: College info, H-I: Gender, J-L: Nationality
        col_college = 5  # Column F (0-indexed, so F=5)
        col_total = 6    # Column G
        col_ladies = 7   # Column H  
        col_gentlemen = 8 # Column I
        col_saudi = 9    # Column J
        col_non_saudi = 10 # Column K
        
        # Write headers
        worksheet.write(0, col_college, "Total of Graduates by Colleges", header_format)
        worksheet.write(0, col_ladies, "Total Graduates by Gender", header_format)
        worksheet.write(0, col_saudi, "Total Graduates by Nationalities", header_format)
        
        # Write sub-headers
        worksheet.write(1, col_ladies, "Ladies", header_format)
        worksheet.write(1, col_gentlemen, "Gentlemen", header_format)
        worksheet.write(1, col_saudi, "Saudi", header_format)
        worksheet.write(1, col_non_saudi, "Non-Saudi", header_format)
        
        # Write data rows
        for idx, row in summary_df.iterrows():
            row_num = idx + 2  # Start from row 2 (0-indexed)
            
            worksheet.write(row_num, col_college, row["College"], number_format)
            worksheet.write(row_num, col_total, row["Total Graduates"], number_format)
            worksheet.write(row_num, col_ladies, row["Ladies"], number_format)
            worksheet.write(row_num, col_gentlemen, row["Gentlemen"], number_format)
            worksheet.write(row_num, col_saudi, row["Saudi"], number_format)
            worksheet.write(row_num, col_non_saudi, row["Non-Saudi"], number_format)
        
        # Write totals row
        total_row = len(summary_df) + 2
        worksheet.write(total_row, col_college, "Total of Graduates", header_format)
        worksheet.write(total_row, col_total, summary_df["Total Graduates"].sum(), number_format)
        worksheet.write(total_row, col_ladies, summary_df["Ladies"].sum(), number_format)
        worksheet.write(total_row, col_gentlemen, summary_df["Gentlemen"].sum(), number_format)
        worksheet.write(total_row, col_saudi, summary_df["Saudi"].sum(), number_format)
        worksheet.write(total_row, col_non_saudi, summary_df["Non-Saudi"].sum(), number_format)
        
        # Set column widths
        worksheet.set_column(col_college, col_college, 35)  # College names
        worksheet.set_column(col_total, col_non_saudi, 15)  # All numeric columns
        
        print("Successfully created All years summary tab")
        
    except Exception as e:
        print(f"Error creating All years summary tab: {str(e)}")


def process_alumni_list(session_id, colleges, years, allowed_statuses, gender_option, nationality_option=None, degree_option="all"):
    """Generate alumni list based on given parameters"""
    try:
        if session_id not in session_data:
            return {"error": "No data found for your session. Please upload an Excel file first."}

        df = session_data[session_id]["data"]
        
        # Clean statuses
        df["Current Status"] = df["Current Status"].apply(clean_status)
        allowed_statuses = [clean_status(status) for status in allowed_statuses]

        # Basic filters
        filtered_df = df[df["College"].str.strip().isin(colleges)]
        filtered_df = filtered_df[filtered_df["Year/Semester of Graduation"].str.strip().isin(years)]

        # Add degree filtering
        if degree_option == "bachelor":
            filtered_df = filtered_df[~filtered_df["Student ID"].str.startswith("G", na=False)]
        elif degree_option == "master":
            filtered_df = filtered_df[filtered_df["Student ID"].str.startswith("G", na=False)]

        # Filter by gender if needed
        if gender_option.lower() != "all":
            filtered_df = filtered_df[filtered_df["Gender"].str.strip().str.lower() == gender_option.lower()]

        # Filter by nationality
        if nationality_option and nationality_option.lower() != "all":
            if nationality_option.lower() == "saudi":
                filtered_df = filtered_df[filtered_df["Nationality"].str.strip() == "Saudi Arabia"]
            elif nationality_option.lower() == "non-saudi":
                filtered_df = filtered_df[filtered_df["Nationality"].str.strip() != "Saudi Arabia"]

        filtered_df = filtered_df[filtered_df["Current Status"].isin(allowed_statuses)]

        if filtered_df.empty:
            return {"error": "No matching data found for the given filters."}

        filtered_df["College Degree"] = filtered_df["College"].str.strip() + " " + \
                                       safe_get_column(filtered_df, "Degree").str.strip()

        # Generate a unique filename
        output_file = f"Alumni_List_{uuid.uuid4().hex[:8]}.xlsx"
        file_path = os.path.join(app.config['GENERATED_FILES'], output_file)
        
        columns_to_output = [
            "Student Name", "Gender", "College Degree",
            "Major", "Minor", "Concentration",
            "Personal Email", "Phone Number", "Nationality",
            "GPA", "Year/Semester of Graduation"
        ]
        
        # Make sure all columns exist or create them with empty values
        for col in columns_to_output:
            if col not in filtered_df.columns:
                filtered_df[col] = ""
                
        df_to_output = filtered_df[columns_to_output]

        with pd.ExcelWriter(file_path, engine="xlsxwriter") as writer:
            college_abbreviations = {
                "College of Engineering & Advan": "CoE",
                "College of Business": "CoB",
                "College of Science & General S": "CoS",
                "College of Medicine": "CoM",
                "College of Pharmacy": "CoP"
            }
            for college in colleges:
                college_df = df_to_output[filtered_df["College"].str.strip() == college]
                if college_df.empty:
                    continue
                sheet_name = college_abbreviations.get(college.strip(), college[:25])
                college_df.to_excel(writer, sheet_name=sheet_name, index=False)

                worksheet = writer.sheets[sheet_name]
                for i, col in enumerate(college_df.columns):
                    series = college_df[col].fillna("")
                    max_len = max(series.astype(str).apply(len).max(), len(col))
                    worksheet.set_column(i, i, max_len + 2)

        return {
            "status": "success",
            "file": output_file
        }
            
    except Exception as e:
        return {"error": f"Error generating alumni list: {str(e)}"}

def process_workplace_report(session_id, colleges, years, degree_option, gender_option, nationality_option=None):
    """Generate workplace statistics report"""
    try:
        if session_id not in session_data:
            return {"error": "No data found for your session. Please upload an Excel file first."}

        df = session_data[session_id]["data"]
        stats = get_workplace_statistics(df, colleges, years, degree_option, gender_option, nationality_option)

        # Generate a unique filename
        output_file = f"Workplace_Report_{uuid.uuid4().hex[:8]}.xlsx"
        file_path = os.path.join(app.config['GENERATED_FILES'], output_file)

        with pd.ExcelWriter(file_path, engine="xlsxwriter") as writer:
            workbook = writer.book

            # Create formats
            header_format = workbook.add_format({
                'bold': True,
                'bg_color': '#4472C4',
                'font_color': 'white',
                'border': 1,
                'align': 'center'
            })
            cell_format = workbook.add_format({
                'border': 1,
                'align': 'left'
            })
            number_format = workbook.add_format({
                'border': 1,
                'align': 'center'
            })

            # Summary Sheet
            summary_df = pd.DataFrame({
                'Metric': [
                    'Total Alumni',
                    'Valid Workplace Entries',
                    'Empty/Unknown Entries',
                    'Number of Colleges',
                    'Years Covered',
                    'Degree Level'
                ],
                'Value': [
                    stats['total_alumni'],
                    stats['valid_entries'],
                    stats['empty_entries'],
                    len(colleges),
                    f"{len(years)} years",
                    degree_option.capitalize()
                ]
            })
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            summary_sheet = writer.sheets['Summary']
            summary_sheet.set_column('A:A', 25)
            summary_sheet.set_column('B:B', 40)

            # Empty/Unknown Entries Analysis Sheet
            empty_df = pd.DataFrame.from_dict(stats['empty_stats'], orient='index', columns=['Count'])
            empty_df.index.name = 'Category'
            empty_df.to_excel(writer, sheet_name='Empty Analysis')
            empty_sheet = writer.sheets['Empty Analysis']
            empty_sheet.set_column('A:A', 40)
            empty_sheet.set_column('B:B', 15)

            # Top Employers Sheet
            employers_df = pd.DataFrame.from_dict(stats['top_employers'], orient='index', columns=['Count'])
            employers_df.index.name = 'Employer'
            employers_df.to_excel(writer, sheet_name='Top Employers')
            employer_sheet = writer.sheets['Top Employers']
            employer_sheet.set_column('A:A', 40)
            employer_sheet.set_column('B:B', 15)

            # Top Positions Sheet (replaced with High Positions)
            high_positions_df = pd.DataFrame.from_dict(stats['high_positions'], orient='index', columns=['Count'])
            high_positions_df.index.name = 'High Position'
            high_positions_df.to_excel(writer, sheet_name='Top Positions')
            position_sheet = writer.sheets['Top Positions']
            position_sheet.set_column('A:A', 40)
            position_sheet.set_column('B:B', 15)
            
            # Add summary of high positions at the top
            position_sheet.write(0, 3, 'Total Alumni with High Positions:')
            position_sheet.write(0, 4, stats['high_positions_count'])
            position_sheet.write(1, 3, '% of Alumni with High Positions:')
            high_position_percentage = (stats['high_positions_count'] / stats['total_alumni'] * 100) if stats['total_alumni'] > 0 else 0
            position_sheet.write(1, 4, f'{high_position_percentage:.2f}%')

            # Industry Distribution Sheet
            industry_df = pd.DataFrame.from_dict(stats['industry_dist'], orient='index', columns=['Count'])
            industry_df.index.name = 'Industry'
            industry_df.to_excel(writer, sheet_name='Industry Distribution')
            industry_sheet = writer.sheets['Industry Distribution']
            industry_sheet.set_column('A:A', 40)
            industry_sheet.set_column('B:B', 15)

        return {
            "status": "success",
            "file": output_file
        }
            
    except Exception as e:
        return {"error": f"Error generating workplace report: {str(e)}"}

def process_banner_integration(banner_session_id, alumni_session_id):
    """Compare Banner graduates with Alumni List and add new graduates from Banner to Alumni List"""
    try:
        if banner_session_id not in session_data:
            return {"error": "No Banner data found for your session. Please upload the Banner Excel file first."}
        
        if alumni_session_id not in session_data:
            return {"error": "No Alumni List data found for your session. Please upload the Alumni List Excel file first."}
        
        # Get both dataframes
        banner_df = session_data[banner_session_id]["data"]
        alumni_df = session_data[alumni_session_id]["data"]
        alumni_file = session_data[alumni_session_id]["file_name"]
        
        # Clean and prepare dataframes
        # Ensure the Student ID column exists in both dataframes
        if "Student ID" not in banner_df.columns:
            return {"error": "Student ID column not found in Banner data. Please check your file format."}
        
        if "Student ID" not in alumni_df.columns:
            return {"error": "Student ID column not found in Alumni List data. Please check your file format."}
        
        # Convert Student IDs to string for comparison
        banner_df["Student ID"] = banner_df["Student ID"].astype(str).str.strip()
        alumni_df["Student ID"] = alumni_df["Student ID"].astype(str).str.strip()
        
        # Find students in Banner who are not in the Alumni List
        banner_student_ids = set(banner_df["Student ID"].tolist())
        alumni_student_ids = set(alumni_df["Student ID"].tolist())
        
        new_student_ids = banner_student_ids - alumni_student_ids
        
        if not new_student_ids:
            return {
                "status": "success",
                "message": "No new graduates found in Banner that are not already in the Alumni List.",
                "new_records": 0,
                "banner_records": len(banner_df),
                "alumni_records": len(alumni_df),
                "new_students": []
            }
        
        # Get the new students from Banner
        new_students_df = banner_df[banner_df["Student ID"].isin(new_student_ids)].copy()
        
        # Field mapping from Banner to Alumni List
        field_mapping = {
            "Graduation Term": "Year/Semester of Graduation",
            "Student ID": "Student ID",
            "Student Name": "Student Name",
            "College": "College",
            "Degree": "Degree",
            "Major": "Major",
            "Minor": "Minor",
            "Concentration": "Concentration",
            "Nationality": "Nationality",
            "SSN": "SSN",
            "Gender": "Gender",
            "Alfaisal Email": "Alfaisal Email", 
            "Personal Email": "Personal Email",
            "Phone Number": "Phone Number",
            "Joined AU": "Joining date",
            "CGPA": "GPA"
        }
        
        # Create a new dataframe with Alumni List structure
        new_alumni_records = []
        
        # Current date for comments
        current_date = datetime.now().strftime("%Y-%m-%d")
        
        # Get the column order from the original alumni file
        alumni_columns = alumni_df.columns.tolist()
        
        for _, row in new_students_df.iterrows():
            new_record = {col: "" for col in alumni_columns}  # Initialize with empty values for all columns
            
            # Map fields from Banner to Alumni List
            for banner_field, alumni_field in field_mapping.items():
                if banner_field in row and alumni_field in alumni_columns:
                    new_record[alumni_field] = row[banner_field]
            
            # Set default values for fields not in Banner
            new_record["Current Status"] = "New graduate"
            new_record["Current Workplace"] = ""
            new_record["Current Position"] = ""
            new_record["Work Industry"] = ""
            new_record["Workplace Country"] = ""
            new_record["Workplace City"] = ""
            new_record["Work Email"] = ""
            new_record["Work Phone"] = ""
            new_record["Social Media and Events"] = ""
            new_record["Comments"] = f"Added from Banner on {current_date}"
            
            new_alumni_records.append(new_record)
        
        # Convert to DataFrame with the same column order
        new_alumni_df = pd.DataFrame(new_alumni_records, columns=alumni_columns)
        
        # Combine the original alumni data with new records
        combined_df = pd.concat([alumni_df, new_alumni_df], ignore_index=True)
        
        # Generate output filename based on the original
        base_name = os.path.splitext(alumni_file)[0]
        output_file = f"{base_name}_Updated_{uuid.uuid4().hex[:8]}.xlsx"
        file_path = os.path.join(app.config['GENERATED_FILES'], output_file)
        
        # Copy the original Excel file to preserve formatting
        original_file = os.path.join(app.config['UPLOAD_FOLDER'], alumni_file)
        shutil.copy2(original_file, file_path)
        
        # Load the workbook to get the table name and style info
        template_wb = load_workbook(file_path)
        template_ws = template_wb.active
        
        # Get table info if it exists
        table_name = None
        table_style_info = None
        if template_ws.tables:
            table = list(template_ws.tables.values())[0]
            table_name = list(template_ws.tables.keys())[0]
            table_style_info = {
                'name': table_name,
                'style': table.tableStyleInfo
            }
        template_wb.close()
        
        # Write the combined data to the new file
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='w') as writer:
            # Write the DataFrame
            combined_df.to_excel(writer, index=False, sheet_name='Sheet1')
            
            # Access the workbook and worksheet
            wb = writer.book
            ws = wb['Sheet1']
            
            # If we had a table before, recreate it with the same style
            if table_style_info:
                from openpyxl.worksheet.table import Table, TableStyleInfo
                tab = Table(displayName=table_style_info['name'],
                          ref=f"A1:{get_column_letter(len(alumni_columns))}{len(combined_df) + 1}")
                
                # Set the table style if it was present in the original
                if table_style_info['style']:
                    style = table_style_info['style']
                    tab.tableStyleInfo = TableStyleInfo(
                        name=style.name,
                        showFirstColumn=style.showFirstColumn,
                        showLastColumn=style.showLastColumn,
                        showRowStripes=style.showRowStripes,
                        showColumnStripes=style.showColumnStripes
                    )
                
                ws.add_table(tab)
        
        # Get list of new students with names for display
        new_students_info = new_students_df[["Student ID", "Student Name"]].values.tolist()
        
        return {
            "status": "success",
            "file": output_file,
            "new_records": len(new_alumni_records),
            "banner_records": len(banner_df),
            "alumni_records": len(alumni_df),
            "combined_records": len(combined_df),
            "new_students": new_students_info
        }
            
    except Exception as e:
        return {"error": f"Error processing Banner integration: {str(e)}"}

# --------------------
# Routes
# --------------------
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({"error": "No file part"})
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({"error": "No selected file"})
    
    if file and file.filename.endswith(('.xls', '.xlsx')):
        # Create a session ID for this upload
        session_id = uuid.uuid4().hex
        
        # Save the file
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{session_id}_{filename}")
        file.save(file_path)
        
        # Process the file
        result = load_excel_data(file_path, session_id)
        
        # Add the session ID to the result
        if isinstance(result, dict) and "error" not in result:
            result["session_id"] = session_id
            
        return jsonify(result)
    
    return jsonify({"error": "Invalid file format. Please upload an Excel file (.xls or .xlsx)"})

@app.route('/get_constants')
def get_constants():
    session_id = request.args.get('session_id')
    years = graduation_years
    
    # If we have a session ID and it has custom years, use those
    if session_id and session_id in session_data:
        if "graduation_years" in session_data[session_id]:
            years = session_data[session_id]["graduation_years"]
    
    # If no years found (either globally or in session), use defaults
    if not years:
        years = default_graduation_years
    
    return jsonify({
        "colleges": college_options,
        "years": years
    })

@app.route('/qaa_preview', methods=['POST'])
def qaa_preview():
    data = request.json
    session_id = data.get('session_id')
    
    if session_id not in session_data:
        return jsonify({"error": "Session not found. Please upload your file again."})
    
    df = session_data[session_id]["data"]
    
    colleges = data.get('colleges', [])
    years = data.get('years', [])
    degree_option = data.get('degree_option', 'all')
    gender_option = data.get('gender_option', 'all')
    nationality_option = data.get('nationality_option', 'all')
    mode_option = data.get('mode_option', 'detailed')
    
    # Filter data
    filtered_df = df[df["_College"].isin(colleges) & df["_Year"].isin(years)]
    
    # Apply gender filter
    if gender_option.lower() != "all":
        filtered_df = filtered_df[filtered_df["Gender"].str.strip().str.lower() == gender_option.lower()]
    
    # Apply nationality filter
    if nationality_option and nationality_option.lower() != "all":
        if nationality_option.lower() == "saudi":
            filtered_df = filtered_df[filtered_df["Nationality"].str.strip() == "Saudi Arabia"]
        elif nationality_option.lower() == "non-saudi":
            filtered_df = filtered_df[filtered_df["Nationality"].str.strip() != "Saudi Arabia"]
    
    # Apply degree filter
    if degree_option == "bachelor":
        filtered_df = filtered_df[~filtered_df["Student ID"].str.startswith("G", na=False)]
    elif degree_option == "master":
        filtered_df = filtered_df[filtered_df["Student ID"].str.startswith("G", na=False)]
    
    # Preview summary
    college_counts = filtered_df["_College"].value_counts().to_dict()
    
    if mode_option == "simple":
        # For Simple mode, show gender breakdown instead of status
        gender_counts = filtered_df["Gender"].value_counts().to_dict()
        return jsonify({
            "total_records": len(filtered_df),
            "college_counts": college_counts,
            "gender_counts": gender_counts,
            "mode": "simple"
        })
    else:
        # For Detailed mode, show status breakdown
        status_counts = filtered_df["_CurrentStatus"].value_counts().head(10).to_dict()
        return jsonify({
            "total_records": len(filtered_df),
            "college_counts": college_counts,
            "status_counts": status_counts,
            "mode": "detailed"
        })

@app.route('/alumni_list_preview', methods=['POST'])
def alumni_list_preview():
    data = request.json
    session_id = data.get('session_id')
    
    if session_id not in session_data:
        return jsonify({"error": "Session not found. Please upload your file again."})
    
    df = session_data[session_id]["data"]
    
    colleges = data.get('colleges', [])
    years = data.get('years', [])
    degree_option = data.get('degree_option', 'all')
    gender_option = data.get('gender_option', 'all')
    nationality_option = data.get('nationality_option', 'all')
    status_options = data.get('status_options', [])
    
    # Filter data
    filtered_df = df[df["_College"].isin(colleges) & df["_Year"].isin(years)]
    
    # Apply degree filter
    if degree_option == "bachelor":
        filtered_df = filtered_df[~filtered_df["Student ID"].str.startswith("G", na=False)]
    elif degree_option == "master":
        filtered_df = filtered_df[filtered_df["Student ID"].str.startswith("G", na=False)]

    # Filter by gender if needed
    if gender_option.lower() != "all":
        filtered_df = filtered_df[filtered_df["Gender"].str.strip().str.lower() == gender_option.lower()]

    # Filter by nationality
    if nationality_option and nationality_option.lower() != "all":
        if nationality_option.lower() == "saudi":
            filtered_df = filtered_df[filtered_df["Nationality"].str.strip() == "Saudi Arabia"]
        elif nationality_option.lower() == "non-saudi":
            filtered_df = filtered_df[filtered_df["Nationality"].str.strip() != "Saudi Arabia"]

    # Filter by status
    filtered_df = filtered_df[filtered_df["_CurrentStatus"].isin(status_options)]
    
    # Preview summary
    college_counts = filtered_df["_College"].value_counts().to_dict()
    gender_counts = filtered_df["Gender"].value_counts().to_dict()
    
    return jsonify({
        "total_alumni": len(filtered_df),
        "college_counts": college_counts,
        "gender_counts": gender_counts
    })

@app.route('/workplace_preview', methods=['POST'])
def workplace_preview():
    data = request.json
    session_id = data.get('session_id')
    
    if session_id not in session_data:
        return jsonify({"error": "Session not found. Please upload your file again."})
    
    df = session_data[session_id]["data"]
    
    colleges = data.get('colleges', [])
    years = data.get('years', [])
    degree_option = data.get('degree_option', 'all')
    gender_option = data.get('gender_option', 'all')
    nationality_option = data.get('nationality_option', 'all')
    
    # Get workplace statistics
    stats = get_workplace_statistics(df, colleges, years, degree_option, gender_option, nationality_option)
    
    # Limit the size of the response for preview
    preview_stats = {
        "total_alumni": stats["total_alumni"],
        "valid_entries": stats["valid_entries"],
        "empty_entries": stats["empty_entries"],
        "empty_stats": {k: stats["empty_stats"][k] for k in list(stats["empty_stats"].keys())[:5]} if stats["empty_stats"] else {},
        "top_employers": {k: stats["top_employers"][k] for k in list(stats["top_employers"].keys())[:5]} if stats["top_employers"] else {},
        "top_positions": {k: stats["high_positions"][k] for k in list(stats["high_positions"].keys())[:5]} if stats["high_positions"] else {},
        "industry_dist": {k: stats["industry_dist"][k] for k in list(stats["industry_dist"].keys())[:3]} if stats["industry_dist"] else {}
    }
    
    return jsonify(preview_stats)

@app.route('/generate_qaa_report', methods=['POST'])
def generate_qaa_report():
    data = request.json
    session_id = data.get('session_id')
    colleges = data.get('colleges', [])
    years = data.get('years', [])
    degree_option = data.get('degree_option', 'all')
    combine_all = data.get('combine_all', False)
    combine_years = data.get('combine_years', True)
    gender_option = data.get('gender_option', 'all')
    nationality_option = data.get('nationality_option', 'all')
    mode_option = data.get('mode_option', 'detailed')
    
    result = process_qaa_report(
        session_id, colleges, years, degree_option, 
        combine_all, combine_years, gender_option, nationality_option, mode_option
    )
    
    return jsonify(result)

@app.route('/generate_alumni_list', methods=['POST'])
def generate_alumni_list():
    data = request.json
    session_id = data.get('session_id')
    colleges = data.get('colleges', [])
    years = data.get('years', [])
    status_options = data.get('status_options', [])
    degree_option = data.get('degree_option', 'all')
    gender_option = data.get('gender_option', 'all')
    nationality_option = data.get('nationality_option', 'all')
    
    result = process_alumni_list(
        session_id, colleges, years, status_options, 
        gender_option, nationality_option, degree_option
    )
    
    return jsonify(result)

@app.route('/generate_workplace_report', methods=['POST'])
def generate_workplace_report():
    data = request.json
    session_id = data.get('session_id')
    colleges = data.get('colleges', [])
    years = data.get('years', [])
    degree_option = data.get('degree_option', 'all')
    gender_option = data.get('gender_option', 'all')
    nationality_option = data.get('nationality_option', 'all')
    
    result = process_workplace_report(
        session_id, colleges, years, degree_option, 
        gender_option, nationality_option
    )
    
    return jsonify(result)

@app.route('/generate_banner_integration', methods=['POST'])
def generate_banner_integration():
    data = request.json
    banner_session_id = data.get('banner_session_id')
    alumni_session_id = data.get('alumni_session_id')
    
    result = process_banner_integration(banner_session_id, alumni_session_id)
    
    return jsonify(result)

@app.route('/download/<filename>')
def download_file(filename):
    return send_from_directory(app.config['GENERATED_FILES'], filename, as_attachment=True)

# Cleanup old files - could be run periodically in production
@app.route('/cleanup', methods=['POST'])
def cleanup_files():
    try:
        # Get list of files in upload and generated folders
        upload_files = os.listdir(app.config['UPLOAD_FOLDER'])
        generated_files = os.listdir(app.config['GENERATED_FILES'])
        
        # Get current time
        now = datetime.now()
        
        # Delete files older than 24 hours
        for file in upload_files:
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], file)
            file_created = datetime.fromtimestamp(os.path.getctime(file_path))
            if (now - file_created).days > 0:  # Older than 1 day
                os.remove(file_path)
                
        for file in generated_files:
            file_path = os.path.join(app.config['GENERATED_FILES'], file)
            file_created = datetime.fromtimestamp(os.path.getctime(file_path))
            if (now - file_created).days > 0:  # Older than 1 day
                os.remove(file_path)
                
        # Clean up old sessions
        sessions_to_remove = []
        for sid, data in session_data.items():
            timestamp = datetime.fromisoformat(data['timestamp'])
            if (now - timestamp).days > 0:  # Older than 1 day
                sessions_to_remove.append(sid)
                
        for sid in sessions_to_remove:
            del session_data[sid]
            
        return jsonify({"status": "success", "message": "Cleanup completed"})
    except Exception as e:
        return jsonify({"error": f"Cleanup failed: {str(e)}"})

if __name__ == '__main__':
    app.run(debug=True)