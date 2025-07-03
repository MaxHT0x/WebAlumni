#!/usr/bin/env python3
"""
Test validator module for QAA reports
Parses generated Excel files and validates against expected results
This module is separate from report generation logic to ensure true testing
"""

import os
import json
from openpyxl import load_workbook
from typing import Dict, List, Union, Any

class QAATestValidator:
    """Validates QAA report Excel files against expected results"""
    
    def __init__(self, expectations_file: str = "/home/rakanlinux/coolProjects/WebAlumni/test_expectations.json"):
        """Initialize validator with expected results"""
        self.expectations_file = expectations_file
        self.expectations = self._load_expectations()
    
    def _load_expectations(self) -> Dict:
        """Load expected results from JSON file"""
        try:
            with open(self.expectations_file, 'r') as f:
                return json.load(f)
        except FileNotFoundError:
            # Return default expectations if file doesn't exist yet
            return {
                "detailed_mode": {
                    "2014-2015": {
                        "total_graduates": 251
                    }
                },
                "simple_mode": {
                    "2014-2015": {
                        "total_graduates": 251,
                        "gentlemen": 162,
                        "ladies": 89
                    }
                }
            }
    
    def validate_detailed_mode_excel(self, excel_file_path: str, test_year: str = "2014-2015") -> Dict:
        """
        Validate detailed mode Excel file
        For detailed mode with 'combine all', we need to find the Total column and sum it
        """
        results = {
            "mode": "detailed",
            "test_year": test_year,
            "passed": False,
            "details": [],
            "errors": []
        }
        
        try:
            if not os.path.exists(excel_file_path):
                results["errors"].append(f"Excel file not found: {excel_file_path}")
                return results
            
            workbook = load_workbook(excel_file_path, data_only=True)
            
            # Look for the main data sheet (could be named differently)
            # Common names might be "Combined Report", "QAA Report", or the first sheet
            data_sheet = None
            possible_sheet_names = ["Combined Report", "QAA Report", workbook.sheetnames[0]]
            
            for sheet_name in possible_sheet_names:
                if sheet_name in workbook.sheetnames:
                    data_sheet = workbook[sheet_name]
                    break
            
            if data_sheet is None:
                results["errors"].append(f"Could not find data sheet. Available sheets: {workbook.sheetnames}")
                return results
            
            # Find the "Total" column
            total_column_index = None
            header_row = 1  # Assume first row is header
            
            for col in range(1, data_sheet.max_column + 1):
                cell_value = data_sheet.cell(row=header_row, column=col).value
                if cell_value and str(cell_value).strip().lower() == "total":
                    total_column_index = col
                    break
            
            if total_column_index is None:
                results["errors"].append("Could not find 'Total' column in Excel sheet")
                return results
            
            # Look for the "Overall Total" row instead of summing all values
            total_graduates = 0
            for row in range(header_row + 1, data_sheet.max_row + 1):
                # Check the first column for "Overall Total" label
                label_cell = data_sheet.cell(row=row, column=1).value
                if label_cell and isinstance(label_cell, str) and "overall total" in label_cell.lower():
                    cell_value = data_sheet.cell(row=row, column=total_column_index).value
                    if cell_value and isinstance(cell_value, (int, float)):
                        total_graduates = int(cell_value)
                        break
            
            # Get expected results
            expected = self.expectations.get("detailed_mode", {}).get(test_year, {})
            expected_total = expected.get("total_graduates", 0)
            
            # Validate
            if total_graduates == expected_total:
                results["passed"] = True
                results["details"].append(f"✅ Total graduates: {total_graduates} (expected: {expected_total})")
            else:
                results["details"].append(f"❌ Total graduates: {total_graduates} (expected: {expected_total})")
            
            results["actual_total"] = total_graduates
            results["expected_total"] = expected_total
            
        except Exception as e:
            results["errors"].append(f"Error validating detailed mode Excel: {str(e)}")
        
        return results
    
    def validate_simple_mode_excel(self, excel_file_path: str, test_year: str = "2014-2015") -> Dict:
        """
        Validate simple mode Excel file
        Simple mode creates sheets by academic year with gender breakdown
        """
        results = {
            "mode": "simple",
            "test_year": test_year,
            "passed": False,
            "details": [],
            "errors": []
        }
        
        try:
            if not os.path.exists(excel_file_path):
                results["errors"].append(f"Excel file not found: {excel_file_path}")
                return results
            
            workbook = load_workbook(excel_file_path, data_only=True)
            
            # Look for the year sheet - could be "2014-2015", "2014_2015", etc.
            year_sheet = None
            possible_sheet_names = [
                test_year,
                test_year.replace("-", "_"),
                test_year.replace("-", ""),
                f"Year_{test_year}",
                f"Academic_{test_year}"
            ]
            
            for sheet_name in possible_sheet_names:
                if sheet_name in workbook.sheetnames:
                    year_sheet = workbook[sheet_name]
                    break
            
            if year_sheet is None:
                results["errors"].append(f"Could not find year sheet for {test_year}. Available sheets: {workbook.sheetnames}")
                return results
            
            # In Simple mode, we need to find gender totals
            # The structure typically has colleges listed with Gentlemen/Ladies columns
            gentlemen_total = 0
            ladies_total = 0
            
            # Look for "Gentlemen" and "Ladies" columns
            gentlemen_col = None
            ladies_col = None
            
            # Search through the first few rows to find headers
            for row in range(1, 6):  # Check first 5 rows for headers
                for col in range(1, year_sheet.max_column + 1):
                    cell_value = year_sheet.cell(row=row, column=col).value
                    if cell_value:
                        cell_text = str(cell_value).strip().lower()
                        if "gentlemen" in cell_text:
                            gentlemen_col = col
                        elif "ladies" in cell_text:
                            ladies_col = col
            
            if gentlemen_col is None or ladies_col is None:
                results["errors"].append(f"Could not find 'Gentlemen' or 'Ladies' columns. Searching for alternative patterns...")
                
                # Alternative: look for "Male"/"Female" columns
                for row in range(1, 6):
                    for col in range(1, year_sheet.max_column + 1):
                        cell_value = year_sheet.cell(row=row, column=col).value
                        if cell_value:
                            cell_text = str(cell_value).strip().lower()
                            if "male" in cell_text and "female" not in cell_text:
                                gentlemen_col = col
                            elif "female" in cell_text:
                                ladies_col = col
            
            if gentlemen_col is None or ladies_col is None:
                results["errors"].append("Could not find gender columns in Simple mode Excel")
                return results
            
            # Look for the TOTAL row instead of summing all values
            for row in range(1, year_sheet.max_row + 1):
                # Check the first column for "TOTAL" label
                label_cell = year_sheet.cell(row=row, column=1).value
                if label_cell and isinstance(label_cell, str) and label_cell.strip().upper() == "TOTAL":
                    gentlemen_value = year_sheet.cell(row=row, column=gentlemen_col).value
                    ladies_value = year_sheet.cell(row=row, column=ladies_col).value
                    
                    if gentlemen_value and isinstance(gentlemen_value, (int, float)):
                        gentlemen_total = int(gentlemen_value)
                    
                    if ladies_value and isinstance(ladies_value, (int, float)):
                        ladies_total = int(ladies_value)
                    break
            
            total_graduates = gentlemen_total + ladies_total
            
            # Get expected results
            expected = self.expectations.get("simple_mode", {}).get(test_year, {})
            expected_total = expected.get("total_graduates", 0)
            expected_gentlemen = expected.get("gentlemen", 0)
            expected_ladies = expected.get("ladies", 0)
            
            # Validate each metric
            all_passed = True
            
            if total_graduates == expected_total:
                results["details"].append(f"✅ Total graduates: {total_graduates} (expected: {expected_total})")
            else:
                results["details"].append(f"❌ Total graduates: {total_graduates} (expected: {expected_total})")
                all_passed = False
            
            if gentlemen_total == expected_gentlemen:
                results["details"].append(f"✅ Gentlemen: {gentlemen_total} (expected: {expected_gentlemen})")
            else:
                results["details"].append(f"❌ Gentlemen: {gentlemen_total} (expected: {expected_gentlemen})")
                all_passed = False
            
            if ladies_total == expected_ladies:
                results["details"].append(f"✅ Ladies: {ladies_total} (expected: {expected_ladies})")
            else:
                results["details"].append(f"❌ Ladies: {ladies_total} (expected: {expected_ladies})")
                all_passed = False
            
            results["passed"] = all_passed
            results["actual_total"] = total_graduates
            results["actual_gentlemen"] = gentlemen_total
            results["actual_ladies"] = ladies_total
            results["expected_total"] = expected_total
            results["expected_gentlemen"] = expected_gentlemen
            results["expected_ladies"] = expected_ladies
            
        except Exception as e:
            results["errors"].append(f"Error validating simple mode Excel: {str(e)}")
        
        return results
    
    def run_full_test_suite(self, simple_mode_file: str, detailed_mode_file: str, test_year: str = "2014-2015") -> Dict:
        """
        Run complete test suite on both Simple and Detailed mode files
        """
        results = {
            "test_year": test_year,
            "overall_passed": False,
            "simple_mode": {},
            "detailed_mode": {},
            "summary": []
        }
        
        # Test Simple mode
        simple_results = self.validate_simple_mode_excel(simple_mode_file, test_year)
        results["simple_mode"] = simple_results
        
        # Test Detailed mode  
        detailed_results = self.validate_detailed_mode_excel(detailed_mode_file, test_year)
        results["detailed_mode"] = detailed_results
        
        # Overall summary
        simple_passed = simple_results.get("passed", False)
        detailed_passed = detailed_results.get("passed", False)
        
        results["overall_passed"] = simple_passed and detailed_passed
        
        if simple_passed:
            results["summary"].append("✅ Simple mode tests PASSED")
        else:
            results["summary"].append("❌ Simple mode tests FAILED")
            
        if detailed_passed:
            results["summary"].append("✅ Detailed mode tests PASSED")
        else:
            results["summary"].append("❌ Detailed mode tests FAILED")
        
        return results

def format_test_results_for_display(results: Dict) -> str:
    """Format test results for web interface display"""
    output = []
    output.append(f"🧪 QAA Report Test Results for {results.get('test_year', 'Unknown Year')}")
    output.append("=" * 50)
    
    # Simple mode results
    simple = results.get("simple_mode", {})
    output.append("\n📊 SIMPLE MODE:")
    if simple.get("errors"):
        for error in simple["errors"]:
            output.append(f"❌ ERROR: {error}")
    else:
        for detail in simple.get("details", []):
            output.append(f"   {detail}")
    
    # Detailed mode results
    detailed = results.get("detailed_mode", {})
    output.append("\n📈 DETAILED MODE:")
    if detailed.get("errors"):
        for error in detailed["errors"]:
            output.append(f"❌ ERROR: {error}")
    else:
        for detail in detailed.get("details", []):
            output.append(f"   {detail}")
    
    # Overall summary
    output.append("\n🎯 SUMMARY:")
    for summary_item in results.get("summary", []):
        output.append(f"   {summary_item}")
    
    if results.get("overall_passed", False):
        output.append("\n🎉 ALL TESTS PASSED! 🎉")
    else:
        output.append("\n⚠️  SOME TESTS FAILED ⚠️")
    
    return "\n".join(output)

# Example usage and testing
if __name__ == "__main__":
    # For testing the validator independently
    validator = QAATestValidator()
    print("Test validator initialized successfully")
    print(f"Expectations loaded: {validator.expectations}")